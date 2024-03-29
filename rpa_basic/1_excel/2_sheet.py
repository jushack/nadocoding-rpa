from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet를 기본 이름으로 생성
ws.title = "mySheet"    # sheet 이름 변경

ws.sheet_properties.tabColor = "ff66ff"   #RGB 형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("yourSheet")  # 주어진 이름으로 sheet 생성됨
ws2 = wb.create_sheet("newSheet", 2)  # 2번째 index에 sheet 생성
new_ws = wb["newSheet"] # dict 형태로 sheet에 접근

print(wb.sheetnames)    # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test"   # 엑셀 A1 위치에 Test 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")